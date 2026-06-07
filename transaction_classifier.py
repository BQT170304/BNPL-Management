"""
Phân loại giao dịch ngân hàng theo từng khách hàng (CIF).

Quy tắc phân loại (3 nhóm loại trừ nhau):
  - Income       : AMOUNT > 0
  - Debt payment : AMOUNT < 0  VÀ  NOTE khớp từ khóa nợ (tra no, tra gop, the tin dung...)
  - Expense      : AMOUNT < 0  (phần còn lại)
  - Uncategorized: AMOUNT == 0 (không thuộc quy tắc nào)

Lưu ý về dữ liệu nguồn:
  File "*.csv.xlsx" tuy là .xlsx nhưng chỉ có 1 cột A, mỗi ô là một dòng CSV
  dạng "CIF_NO,NOTE,TRAN_DATE,AMOUNT". Lớp này tự xử lý cả trường hợp đó lẫn
  file .csv thông thường.
"""
from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd


class BankTransactionClassifier:
    # --- Tên nhãn (constants, tránh magic strings) -------------------------
    INCOME = "Income"
    EXPENSE = "Expense"
    DEBT = "Debt payment"
    UNCATEGORIZED = "Uncategorized"

    # Cột chuẩn của dữ liệu
    COLUMNS = ["CIF_NO", "NOTE", "TRAN_DATE", "AMOUNT"]

    # --- Từ khóa nhận diện nghĩa vụ trả nợ ---------------------------------
    # So khớp trên NOTE đã chuẩn hóa (chữ thường, bỏ dấu tiếng Việt).
    # Dùng ranh giới từ (\b) để hạn chế khớp nhầm (vd "tra nghiep" KHÔNG khớp "tra no").
    DEFAULT_DEBT_KEYWORDS = [
        "tra no",          # tra no the, tra no thau chi, tra no...
        "tra gop",         # tra gop thang, dang ky tra gop
        "the tin dung",    # thanh toan the tin dung
        "tt the",          # tt the = thanh toan the (thẻ)
        "thanh toan the",  # thanh toan the
        "thau chi",        # nợ thấu chi
        "tra cham",        # mua trả chậm
        "khoan vay",
        "lai vay",
        "tra lai",         # trả lãi
        "vay",             # khoản vay / giải ngân vay
        "no vay",
    ]

    def __init__(self, debt_keywords: list[str] | None = None) -> None:
        keywords = debt_keywords if debt_keywords is not None else self.DEFAULT_DEBT_KEYWORDS
        # Một regex duy nhất, có ranh giới từ, không phân biệt hoa thường.
        pattern = "|".join(rf"\b{re.escape(kw)}\b" for kw in keywords)
        self._debt_re = re.compile(pattern)

    # ======================================================================
    # 1) ĐỌC DỮ LIỆU
    # ======================================================================
    def load(self, path: str | Path) -> pd.DataFrame:
        """Đọc file nguồn -> DataFrame[CIF_NO, NOTE, TRAN_DATE, AMOUNT]."""
        path = Path(path)
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            lines = self._read_single_column_xlsx(path)
        else:  # .csv hoặc .txt
            lines = path.read_text(encoding="utf-8-sig").splitlines()
        return self._parse_lines(lines)

    @staticmethod
    def _read_single_column_xlsx(path: Path) -> list[str]:
        """Lấy toàn bộ giá trị cột A (mỗi ô = 1 dòng CSV) từ sheet đầu tiên."""
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # sheet "banking_simulation_6M"
        lines: list[str] = []
        for (value,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if value is not None:
                lines.append(str(value))
        wb.close()
        return lines

    @classmethod
    def _parse_lines(cls, lines: list[str]) -> pd.DataFrame:
        """
        Tách mỗi dòng "CIF_NO,NOTE,TRAN_DATE,AMOUNT".
        Parse an toàn với NOTE có thể chứa dấu phẩy:
          - CIF_NO  = trước dấu phẩy đầu tiên
          - AMOUNT  = sau dấu phẩy cuối cùng
          - TRAN_DATE = field trước AMOUNT
          - NOTE    = phần còn lại ở giữa
        """
        records: list[tuple[str, str, str, str]] = []
        for raw in lines:
            line = raw.strip()
            if not line:
                continue
            # Bỏ qua dòng header nếu gặp
            if line.replace(" ", "").upper().startswith("CIF_NO,NOTE"):
                continue
            try:
                cif, rest = line.split(",", 1)
                note, tran_date, amount = rest.rsplit(",", 2)
            except ValueError:
                continue  # dòng dị dạng -> bỏ
            records.append((cif.strip(), note.strip(), tran_date.strip(), amount.strip()))

        df = pd.DataFrame.from_records(records, columns=cls.COLUMNS)
        df["AMOUNT"] = pd.to_numeric(df["AMOUNT"], errors="coerce").fillna(0)
        df["TRAN_DATE"] = pd.to_datetime(df["TRAN_DATE"], errors="coerce")
        return df

    # ======================================================================
    # 2) CHUẨN HÓA & PHÂN LOẠI
    # ======================================================================
    @staticmethod
    def _normalize(text: str) -> str:
        """Chữ thường, bỏ dấu tiếng Việt, gom khoảng trắng -> dễ so khớp."""
        text = unicodedata.normalize("NFKD", str(text))
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace("đ", "d").replace("Đ", "d")
        return re.sub(r"\s+", " ", text).strip().lower()

    def classify_one(self, note: str, amount: float) -> str:
        """Phân loại 1 giao dịch (dùng cho test / xử lý lẻ)."""
        if amount > 0:
            return self.INCOME
        if amount == 0:
            return self.UNCATEGORIZED
        norm = self._normalize(note)
        return self.DEBT if self._debt_re.search(norm) else self.EXPENSE

    def classify(self, df: pd.DataFrame) -> pd.DataFrame:
        """Thêm cột CATEGORY (vectorized) cho toàn bộ DataFrame."""
        note_norm = df["NOTE"].map(self._normalize)
        is_debt_note = note_norm.str.contains(self._debt_re)
        amount = df["AMOUNT"]

        conditions = [
            amount > 0,
            (amount < 0) & is_debt_note,
            amount < 0,
        ]
        choices = [self.INCOME, self.DEBT, self.EXPENSE]

        out = df.copy()
        out["CATEGORY"] = np.select(conditions, choices, default=self.UNCATEGORIZED)
        return out

    # ======================================================================
    # 3) TỔNG HỢP THEO NHÓM (CIF, hoặc CIF + tháng)
    # ======================================================================
    def _summarize(self, df: pd.DataFrame, group_keys: list[str]) -> pd.DataFrame:
        """
        Hàm tổng hợp dùng chung: 1 dòng / tổ hợp group_keys với tổng tiền
        & số lượng giao dịch theo từng nhóm.
        Income/Expense/Debt để ở giá trị dương (đã lấy trị tuyệt đối).
        net_cashflow = tổng AMOUNT thực = income - expense - debt.
        """
        if "CATEGORY" not in df.columns:
            df = self.classify(df)

        cats = [self.INCOME, self.EXPENSE, self.DEBT]
        sums = (
            df.pivot_table(index=group_keys, columns="CATEGORY",
                           values="AMOUNT", aggfunc="sum", fill_value=0)
            .reindex(columns=cats, fill_value=0)
        )
        counts = (
            df.pivot_table(index=group_keys, columns="CATEGORY",
                           values="AMOUNT", aggfunc="size", fill_value=0)
            .reindex(columns=cats, fill_value=0)
        )

        summary = pd.DataFrame({
            "total_income": sums[self.INCOME],
            "total_expense": sums[self.EXPENSE].abs(),
            "total_debt_payment": sums[self.DEBT].abs(),
            "count_income": counts[self.INCOME].astype(int),
            "count_expense": counts[self.EXPENSE].astype(int),
            "count_debt_payment": counts[self.DEBT].astype(int),
        })
        summary["net_cashflow"] = (
            summary["total_income"] - summary["total_expense"] - summary["total_debt_payment"]
        )
        summary["total_transactions"] = df.groupby(group_keys).size()
        return summary.reset_index().sort_values(group_keys).reset_index(drop=True)

    @staticmethod
    def _add_month(df: pd.DataFrame) -> pd.DataFrame:
        """Thêm cột MONTH dạng 'YYYY-MM' suy ra từ TRAN_DATE."""
        out = df.copy()
        out["MONTH"] = out["TRAN_DATE"].dt.to_period("M").astype(str)
        return out

    def summarize_by_cif(self, df: pd.DataFrame) -> pd.DataFrame:
        """Tổng hợp gộp toàn bộ kỳ -> 1 dòng / CIF."""
        return self._summarize(df, ["CIF_NO"])

    def summarize_by_cif_month(self, df: pd.DataFrame) -> pd.DataFrame:
        """Tổng hợp theo từng tháng -> 1 dòng / (CIF, MONTH)."""
        return self._summarize(self._add_month(df), ["CIF_NO", "MONTH"])

    # ======================================================================
    # 4) PIPELINE TIỆN DỤNG
    # ======================================================================
    def run(self, source: str | Path,
            tx_out: str | Path | None = "transactions_labeled.csv",
            summary_out: str | Path | None = "summary_by_cif.csv",
            monthly_out: str | Path | None = "summary_by_cif_month.csv"
            ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Đọc -> gắn nhãn -> tổng hợp (toàn kỳ + theo tháng) -> (tùy chọn) xuất CSV."""
        df = self.load(source)
        labeled = self.classify(df)
        summary = self.summarize_by_cif(labeled)
        monthly = self.summarize_by_cif_month(labeled)

        if tx_out:
            labeled.to_csv(tx_out, index=False, encoding="utf-8-sig")
        if summary_out:
            summary.to_csv(summary_out, index=False, encoding="utf-8-sig")
        if monthly_out:
            monthly.to_csv(monthly_out, index=False, encoding="utf-8-sig")
        return labeled, summary, monthly


if __name__ == "__main__":
    import sys

    src = sys.argv[1] if len(sys.argv) > 1 else "Dữ liệu giả lập - banking_simulation_6M.csv.xlsx"
    clf = BankTransactionClassifier()
    labeled, summary, monthly = clf.run(src)

    print(f"Tổng giao dịch: {len(labeled):,}")
    print("\nPhân bố theo nhóm:")
    print(labeled["CATEGORY"].value_counts().to_string())
    print(f"\nSố khách hàng (CIF): {summary['CIF_NO'].nunique():,}")
    print(f"Số dòng tổng hợp theo tháng (CIF x MONTH): {len(monthly):,}")
    print("\nTổng hợp toàn kỳ - 3 KH đầu:")
    print(summary.head(3).to_string(index=False))
    print("\nTổng hợp theo tháng - 6 dòng đầu:")
    print(monthly.head(6).to_string(index=False))
    print("\nĐã xuất: transactions_labeled.csv, summary_by_cif.csv, summary_by_cif_month.csv")
